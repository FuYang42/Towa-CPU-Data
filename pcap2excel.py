#!/usr/bin/env python3
"""
PCAP to Excel Converter
一键将 PCAP 文件转换为包含图表的 Excel 文件

使用方法:
    python3 pcap2excel.py <pcap文件路径> [输出excel文件名]

示例:
    python3 pcap2excel.py cpu_usage.pcap
    python3 pcap2excel.py data.pcap output.xlsx
"""

import struct
import sys
import os


def parse_ethernet(data):
    if len(data) < 14:
        return None
    return {'ethertype': struct.unpack('!H', data[12:14])[0], 'payload': data[14:]}


def parse_ip(data):
    if len(data) < 20:
        return None
    ihl = (data[0] & 0xF) * 4
    protocol = data[9]
    return {'protocol': protocol, 'payload': data[ihl:]}


def parse_udp(data):
    if len(data) < 8:
        return None
    payload = data[8:]
    return {'payload': payload, 'payload_length': len(payload)}


def extract_cpu_data(raw_data):
    """提取 CPU 数据"""
    if not raw_data or len(raw_data) < 24:
        return None

    last_24_bytes = raw_data[-24:]
    busy_time_bytes = last_24_bytes[8:16]
    idle_time_bytes = last_24_bytes[16:24]

    try:
        busy_time = struct.unpack('<q', busy_time_bytes)[0]
        idle_time = struct.unpack('<q', idle_time_bytes)[0]
        total_time = busy_time + idle_time

        if total_time > 0:
            cpu_usage = (busy_time / total_time) * 100
            return {
                'cpu_usage': cpu_usage,
                'busy_time': busy_time,
                'idle_time': idle_time,
                'total_time': total_time
            }
    except:
        pass

    return None


def read_pcap(filename, target_payload_length=504, start_index=None, end_index=None):
    """读取 PCAP 文件并提取 CPU 数据

    Args:
        filename: PCAP 文件路径
        target_payload_length: 目标 UDP payload 长度（默认 504）
        start_index: 开始索引（从 1 开始），None 表示从第一个开始
        end_index: 结束索引（包含），None 表示到最后一个
    """
    print(f"正在读取 PCAP 文件: {filename}")
    print(f"只分析 UDP payload 长度为 {target_payload_length} 字节的数据包")

    if start_index is not None or end_index is not None:
        range_str = f"第 {start_index or 1} 个"
        if end_index:
            range_str += f" 到第 {end_index} 个"
        else:
            range_str += " 到最后"
        print(f"分析范围: {range_str}\n")
    else:
        print(f"分析范围: 所有匹配的数据包\n")

    cpu_data = []
    all_valid_data = []  # 先收集所有有效数据

    with open(filename, 'rb') as f:
        # 读取 PCAP 文件头
        header = f.read(24)
        if len(header) < 24:
            raise ValueError("无效的 PCAP 文件")

        magic = struct.unpack('I', header[0:4])[0]
        endian = '<' if magic == 0xa1b2c3d4 else '>'

        # 读取数据包
        packet_count = 0
        while True:
            # 读取包头
            packet_header = f.read(16)
            if len(packet_header) < 16:
                break

            incl_len = struct.unpack(endian + 'I', packet_header[8:12])[0]
            packet_data = f.read(incl_len)
            packet_count += 1

            # 解析数据包
            eth = parse_ethernet(packet_data)
            if eth and eth['ethertype'] == 0x0800:
                ip = parse_ip(eth['payload'])
                if ip and ip['protocol'] == 17:
                    udp = parse_udp(ip['payload'])
                    if udp and udp['payload_length'] == target_payload_length:
                        data = extract_cpu_data(udp['payload'])
                        if data:
                            all_valid_data.append(data)

    # 应用范围过滤
    start = (start_index - 1) if start_index else 0
    end = end_index if end_index else len(all_valid_data)

    cpu_data = all_valid_data[start:end]

    # 重新编号
    for idx, data in enumerate(cpu_data, start=1):
        data['number'] = idx

    print(f"完成！总共处理 {packet_count} 个数据包，找到 {len(all_valid_data)} 个匹配数据包")
    print(f"提取 {len(cpu_data)} 个 CPU 数据点进行分析\n")
    return cpu_data


def export_to_excel(cpu_data, output_file):
    """导出到 Excel（带图表）"""
    try:
        from openpyxl import Workbook
        from openpyxl.chart import LineChart, Reference
        from openpyxl.chart.text import RichText
        from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        print("错误: 需要安装 openpyxl 库")
        print("请运行: pip install openpyxl")
        print("\n或者使用 CSV 模式（无需安装库）:")
        print(f"  python3 pcap2excel.py {sys.argv[1]} --csv")
        sys.exit(1)

    print(f"正在生成 Excel 文件: {output_file}")

    wb = Workbook()
    ws = wb.active
    ws.title = "CPU Usage Data"

    # 设置表头
    headers = ["Number", "CPU Usage%", "Busy Time", "Idle Time", "Busy Time + Idle Time"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 写入数据
    for idx, data in enumerate(cpu_data, start=2):
        ws.cell(row=idx, column=1, value=data['number'])
        ws.cell(row=idx, column=2, value=round(data['cpu_usage'], 2))
        ws.cell(row=idx, column=3, value=data['busy_time'])
        ws.cell(row=idx, column=4, value=data['idle_time'])
        ws.cell(row=idx, column=5, value=data['total_time'])

    # 设置列宽
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 25

    # 分析 CPU 数据，智能确定纵坐标范围
    cpu_values = [d['cpu_usage'] for d in cpu_data]
    min_cpu = min(cpu_values)
    max_cpu = max(cpu_values)
    cpu_range = max_cpu - min_cpu

    # 判断数据模式并设置合适的纵坐标范围
    if cpu_range < 10:
        # 小波动模式（如16%左右波动）：使用更大的纵坐标范围，让线看起来平缓
        y_min = max(0, min_cpu - 15)
        y_max = min(100, max_cpu + 15)
        print(f"检测到小波动模式 (范围: {cpu_range:.2f}%)，使用宽纵坐标范围 [{y_min:.1f}%, {y_max:.1f}%] 以显示平缓趋势")
    else:
        # On-Off 模式或大波动：使用适中的纵坐标范围，体现波动
        margin = cpu_range * 0.2  # 上下各留20%的边距
        y_min = max(0, min_cpu - margin)
        y_max = min(100, max_cpu + margin)
        print(f"检测到波动模式 (范围: {cpu_range:.2f}%)，使用适中纵坐标范围 [{y_min:.1f}%, {y_max:.1f}%] 以体现波动")

    # 创建 CPU 使用率图表
    chart = LineChart()
    chart.title = "CPU Usage Over Time"
    chart.height = 12  # 缩小图表高度
    chart.width = 20   # 缩小图表宽度

    # 隐藏图例
    chart.legend = None

    # 设置纵坐标（Y轴）
    chart.y_axis.title = 'CPU Usage (%)'
    chart.y_axis.scaling.min = y_min
    chart.y_axis.scaling.max = y_max
    chart.y_axis.delete = False
    chart.y_axis.tickLblPos = "nextTo"

    # 增大Y轴刻度数值字体
    chart.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties(sz=1400)))])  # 字号14pt

    # 设置主要网格线间隔
    if cpu_range < 10:
        chart.y_axis.majorUnit = 5  # 小波动时使用5%间隔
    else:
        chart.y_axis.majorUnit = 10  # 大波动时使用10%间隔

    # 显示Y轴刻度数值
    chart.y_axis.numFmt = '0.00'

    # 设置横坐标（X轴）
    chart.x_axis.title = 'Sample Number'
    chart.x_axis.delete = False
    chart.x_axis.tickLblPos = "low"
    chart.x_axis.numFmt = 'General'
    chart.x_axis.tickLblSkip = 5  # 每隔5个显示一次刻度标签
    chart.x_axis.tickMarkSkip = 5  # 每隔5个显示一次刻度线

    # 增大X轴刻度数值字体
    chart.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties(sz=1400)))])  # 字号14pt

    data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(cpu_data) + 1)
    category_ref = Reference(ws, min_col=1, min_row=2, max_row=len(cpu_data) + 1)

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(category_ref)
    chart.series[0].smooth = True
    chart.series[0].graphicalProperties.line.solidFill = "4472C4"
    chart.series[0].graphicalProperties.line.width = 25000  # 加粗线条

    ws.add_chart(chart, "G2")  # 调整位置

    # 创建 Busy/Idle 对比图
    chart2 = LineChart()
    chart2.title = "Busy Time vs Idle Time"
    chart2.height = 12  # 缩小图表高度
    chart2.width = 20   # 缩小图表宽度

    # 设置Y轴
    chart2.y_axis.title = 'Time (ns)'
    chart2.y_axis.delete = False
    chart2.y_axis.tickLblPos = "nextTo"
    chart2.y_axis.numFmt = '#,##0'  # 显示带千位分隔符的整数格式

    # 增大Y轴刻度数值字体
    chart2.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties(sz=1400)))])

    # 设置X轴
    chart2.x_axis.title = 'Sample Number'
    chart2.x_axis.delete = False
    chart2.x_axis.tickLblPos = "low"
    chart2.x_axis.numFmt = 'General'
    chart2.x_axis.tickLblSkip = 5  # 每隔5个显示一次刻度标签
    chart2.x_axis.tickMarkSkip = 5  # 每隔5个显示一次刻度线

    # 增大X轴刻度数值字体
    chart2.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties(sz=1400)))])

    busy_ref = Reference(ws, min_col=3, min_row=1, max_row=len(cpu_data) + 1)
    idle_ref = Reference(ws, min_col=4, min_row=1, max_row=len(cpu_data) + 1)

    chart2.add_data(busy_ref, titles_from_data=True)
    chart2.add_data(idle_ref, titles_from_data=True)
    chart2.set_categories(category_ref)

    chart2.series[0].graphicalProperties.line.solidFill = "ED7D31"
    chart2.series[1].graphicalProperties.line.solidFill = "70AD47"
    chart2.series[0].smooth = True
    chart2.series[1].smooth = True
    chart2.series[0].graphicalProperties.line.width = 25000
    chart2.series[1].graphicalProperties.line.width = 25000

    ws.add_chart(chart2, "G28")  # 调整位置，与第一个图表分开

    # 保存文件
    wb.save(output_file)
    print(f"✓ 成功保存: {output_file}")


def export_to_csv(cpu_data, output_file):
    """导出到 CSV（无需额外库）"""
    print(f"正在生成 CSV 文件: {output_file}")

    with open(output_file, 'w', encoding='utf-8') as f:
        f.write("Number,CPU Usage%,Busy Time,Idle Time,Busy Time + Idle Time\n")
        for data in cpu_data:
            f.write(f"{data['number']},{data['cpu_usage']:.2f},"
                   f"{data['busy_time']},{data['idle_time']},{data['total_time']}\n")

    print(f"✓ 成功保存: {output_file}")
    print("\n提示: 可以用 Excel 打开此 CSV 文件，然后插入图表来可视化数据")


def print_stats(cpu_data):
    """打印统计信息"""
    if not cpu_data:
        print("警告: 没有找到有效的 CPU 数据")
        return

    cpu_values = [d['cpu_usage'] for d in cpu_data]

    print("\nCPU 使用率统计:")
    print(f"  数据点数量: {len(cpu_data)}")
    print(f"  最小值: {min(cpu_values):.2f}%")
    print(f"  最大值: {max(cpu_values):.2f}%")
    print(f"  平均值: {sum(cpu_values) / len(cpu_values):.2f}%")
    print(f"  中位数: {sorted(cpu_values)[len(cpu_values)//2]:.2f}%")


def main():
    if len(sys.argv) < 2:
        print("PCAP to Excel 转换工具")
        print("="*60)
        print("\n用法:")
        print("  python3 pcap2excel.py <pcap文件路径> [输出文件名] [选项]")
        print("\n示例:")
        print("  # 分析所有数据包（默认）")
        print("  python3 pcap2excel.py cpu_usage.pcap")
        print("")
        print("  # 只分析前 3 个数据包")
        print("  python3 pcap2excel.py cpu_usage.pcap --count 3")
        print("")
        print("  # 分析第 2 到第 4 个数据包")
        print("  python3 pcap2excel.py cpu_usage.pcap --range 2 4")
        print("")
        print("  # 指定输出文件名")
        print("  python3 pcap2excel.py data.pcap output.xlsx")
        print("")
        print("  # 导出为 CSV 格式")
        print("  python3 pcap2excel.py data.pcap --csv")
        print("\n选项:")
        print("  --csv              导出为 CSV 格式（无需安装 openpyxl）")
        print("  --count N          只分析前 N 个匹配的数据包")
        print("  --range START END  分析第 START 到第 END 个数据包（包含）")
        sys.exit(1)

    pcap_file = sys.argv[1]

    # 检查文件是否存在
    if not os.path.exists(pcap_file):
        print(f"错误: 文件不存在: {pcap_file}")
        sys.exit(1)

    # 解析参数
    use_csv = '--csv' in sys.argv
    start_index = None
    end_index = None

    # 检查 --count 参数
    if '--count' in sys.argv:
        idx = sys.argv.index('--count')
        if idx + 1 < len(sys.argv):
            try:
                count = int(sys.argv[idx + 1])
                start_index = 1
                end_index = count
            except ValueError:
                print(f"错误: --count 参数必须是数字")
                sys.exit(1)

    # 检查 --range 参数
    if '--range' in sys.argv:
        idx = sys.argv.index('--range')
        if idx + 2 < len(sys.argv):
            try:
                start_index = int(sys.argv[idx + 1])
                end_index = int(sys.argv[idx + 2])
            except ValueError:
                print(f"错误: --range 参数必须是两个数字")
                sys.exit(1)

    # 确定输出文件名
    output_file = None
    for i, arg in enumerate(sys.argv[2:], start=2):
        if not arg.startswith('--') and i > 1:
            # 检查前一个参数是否是选项
            if i > 2 and sys.argv[i-1] in ['--count', '--range']:
                continue
            if i > 3 and sys.argv[i-2] == '--range':
                continue
            output_file = arg
            break

    if not output_file:
        base_name = os.path.splitext(os.path.basename(pcap_file))[0]
        output_file = f"{base_name}_analysis.{'csv' if use_csv else 'xlsx'}"

    # 确保输出文件有正确的扩展名
    if use_csv and not output_file.endswith('.csv'):
        output_file = os.path.splitext(output_file)[0] + '.csv'
    elif not use_csv and not output_file.endswith('.xlsx'):
        output_file = os.path.splitext(output_file)[0] + '.xlsx'

    print("="*60)
    print("PCAP to Excel 转换工具")
    print("="*60)
    print()

    # 读取 PCAP 文件
    try:
        cpu_data = read_pcap(pcap_file, start_index=start_index, end_index=end_index)
    except Exception as e:
        print(f"错误: 无法读取 PCAP 文件: {e}")
        sys.exit(1)

    if not cpu_data:
        print("错误: 没有找到有效的 CPU 数据")
        sys.exit(1)

    # 导出文件
    try:
        if use_csv:
            export_to_csv(cpu_data, output_file)
        else:
            export_to_excel(cpu_data, output_file)
    except Exception as e:
        print(f"错误: 导出失败: {e}")
        sys.exit(1)

    # 打印统计信息
    print_stats(cpu_data)

    print("\n完成！")


if __name__ == "__main__":
    main()
